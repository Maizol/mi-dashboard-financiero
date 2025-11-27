import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, PatternFill, NamedStyle, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO 
import numpy as np
import altair as alt

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(layout="wide", page_title="Dashboard 'El Profe'")

# ===================== INICIO DE LA MEJORA DE DISEÑO (CSS) =====================
st.markdown("""
<style>
    /* Estilos se mantienen... */
    [data-testid="stAppViewContainer"] > .main {
        background-color: #F0F2F5;
    }
    [data-testid="stSidebar"] {
        background-color: #0F172A; 
    }
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3, [data-testid="stSidebar"] p, [data-testid="stSidebar"] label {
        color: white; 
    }
    
    /* ESTILOS PARA ENLACES DE NAVEGACIÓN DE PÁGINAS */
    [data-testid="stSidebar"] a, [data-testid="stSidebar"] a:link, [data-testid="stSidebar"] a:visited {
        color: white !important;
        text-decoration: none;
    }
    [data-testid="stSidebar"] a:hover {
        color: #D6EAF8 !important;
        text-decoration: underline;
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
        color: white !important;
    }
    /* Enlaces de navegación específicos */
    [data-testid="stSidebar"] ul li a {
        color: white !important;
    }
    [data-testid="stSidebar"] .stPageLink {
        color: white !important;
    }
    [data-testid="stVerticalBlock"] [data-testid="stContainer"] {
        background-color: #FFFFFF; 
        border: 1px solid #E0E0E0; 
        box-shadow: 0 4px 8px rgba(0,0,0,0.1); 
        border-radius: 10px; 
        padding: 20px; 
    }
    [data-testid="stVerticalBlock"] [data-testid="stContainer"] h1, [data-testid="stVerticalBlock"] [data-testid="stContainer"] h2, [data-testid="stVerticalBlock"] [data-testid="stContainer"] h3, [data-testid="stVerticalBlock"] [data-testid="stContainer"] h4 {
        color: #0F172A; 
    }
    .st-emotion-cache-z5fcl4 {
        padding-top: 0px;
    }
    .risk-high { background-color: #F8D7DA; color: #721C24; padding: 10px; border-radius: 5px; font-weight: bold;}
    .risk-medium { background-color: #FFF3CD; color: #856404; padding: 10px; border-radius: 5px; font-weight: bold;}
    .risk-low { background-color: #D4EDDA; color: #155724; padding: 10px; border-radius: 5px; font-weight: bold;}
    
</style>
""", unsafe_allow_html=True)
# ======================= FIN DE LA MEJORA DE DISEÑO (CSS) ======================


# --- 1. LÓGICA DE CÁLCULO (FUNCIÓN CORREGIDA Y PRECISA) ---
@st.cache_data 
def calcular_flujo_de_caja(volumen_inicial, crecimiento_mensual, ingreso_ecommerce, inversion_nov, cash_inicial, precio_promedio, dias_cobro):
    
    months = ["Nov-25", "Dic-25", "Ene-26", "Feb-26", "Mar-26", "Abr-26"]
    
    # Parámetros Fijos de Costos
    costo_variable_unidad = 10 
    costo_mano_obra = 500
    costo_alquiler = 800
    costo_servicios = 200
    costo_marketing_base = 300 
    costo_capacitacion = 200
    costo_ecologico = 100
    
    # Cálculo de desfase de cobro en meses
    desfase_meses = max(1, int(np.ceil(dias_cobro / 30))) 
    
    # Costos Fijos para el BEP
    gastos_grales_base_mensual = costo_alquiler + costo_servicios + costo_marketing_base + costo_capacitacion + costo_ecologico
    costos_fijos_operacionales = costo_mano_obra + gastos_grales_base_mensual
    
    # CÁLCULO DEL PUNTO DE EQUILIBRIO MENSUAL (BEP)
    contribucion_marginal_unidad = precio_promedio - costo_variable_unidad
    bep_unidades = costos_fijos_operacionales / contribucion_marginal_unidad if contribucion_marginal_unidad > 0 else 99999
    bep_soles = bep_unidades * precio_promedio

    # --- Generación de Ingresos Proyectados ---
    ingresos_data = []
    volumen_actual = volumen_inicial
    
    # Almacenar ventas de servicios para el cálculo de crédito (necesario para precisión)
    ventas_servicios_historicas = [] 
    
    for i, month in enumerate(months):
        
        if i == 0:
            total_servicios = 1787.50 
            ing_basico = 560.00
            ing_completo = 665.00
            ing_premium = 562.50
        else:
            total_servicios = volumen_actual * precio_promedio
            ing_basico = total_servicios * 0.33 
            ing_completo = total_servicios * 0.33
            ing_premium = total_servicios * 0.33
            
        total_ingresos = total_servicios + ingreso_ecommerce
        
        # Guardar el valor exacto del total de servicios para usarlo en el cálculo de crédito futuro
        ventas_servicios_historicas.append(total_servicios)
        
        ingresos_data.append({
            "Mes": month,
            "Volumen Servicios": volumen_actual,
            "Ingr. Básico (Aprox)": ing_basico,
            "Ingr. Completo (Aprox)": ing_completo,
            "Ingr. Premium (Aprox)": ing_premium,
            "TOTAL SERVICIOS (A)": total_servicios,
            "Ingr. E-commerce (B)": ingreso_ecommerce,
            "TOTAL INGRESOS (A+B)": total_ingresos,
            "Volumen BEP (Unidades)": bep_unidades
        })
        volumen_actual *= (1 + crecimiento_mensual)
        
    df_ingresos = pd.DataFrame(ingresos_data).set_index("Mes")

    # --- Datos de Egresos y Utilidad ---
    egresos_data = []
    for i, month in enumerate(months):
        
        costo_materiales = df_ingresos.iloc[i]["Volumen Servicios"] * costo_variable_unidad 
        costo_marketing = 800 if i == 0 else costo_marketing_base 
        
        egreso_gastos_grales_flujo = costo_alquiler + costo_servicios + costo_marketing + costo_capacitacion + costo_ecologico
        
        total_gastos_fijos_mes = costo_mano_obra + egreso_gastos_grales_flujo
        
        utilidad_bruta = df_ingresos.iloc[i]["TOTAL INGRESOS (A+B)"] - costo_materiales
        utilidad_neta = utilidad_bruta - total_gastos_fijos_mes
        
        egresos_data.append({
            "Mes": month,
            "Costo Mat. (Var)": round(costo_materiales, 2),
            "Costo M.O. (Fijo)": costo_mano_obra,
            "Alquiler": costo_alquiler,
            "Servicios (Luz/Agua)": costo_servicios,
            "Marketing": costo_marketing,
            "Capacitación": costo_capacitacion,
            "Cert. Ecológica": costo_ecologico,
            "Gastos Grales (Fijos)": egreso_gastos_grales_flujo,
            "TOTAL EGRESOS OPERATIVOS": round(costo_materiales + total_gastos_fijos_mes, 2),
            "UTILIDAD NETA OPERACIONAL": round(utilidad_neta, 2)
        })
    df_egresos = pd.DataFrame(egresos_data).set_index("Mes")

    # --- Datos de Flujo de Caja (CÁLCULO DE CRÉDITO DINÁMICO) ---
    flujo_data = []
    saldo_mes_anterior = cash_inicial
    
    for i, month in enumerate(months):
        
        if i == 0:
            ingreso_ventas = (1787.50 * 0.8) + ingreso_ecommerce 
            ingreso_credito = 0 
        else:
            ingreso_ventas = (df_ingresos.iloc[i]["TOTAL SERVICIOS (A)"] * 0.8) + df_ingresos.iloc[i]["Ingr. E-commerce (B)"]
            
            # Aplica el desfase: busca el 20% de las ventas de servicios del mes de origen
            ingreso_credito = 0
            mes_origen_credito = i - desfase_meses 
            if mes_origen_credito >= 0 and mes_origen_credito < len(ventas_servicios_historicas):
                 ingreso_credito = ventas_servicios_historicas[mes_origen_credito] * 0.20
            
        otros_ingresos = 0
        total_ingreso = round(ingreso_ventas + ingreso_credito + otros_ingresos, 2)
        
        egreso_materiales = df_egresos.iloc[i]["Costo Mat. (Var)"]
        egreso_mano_obra = df_egresos.iloc[i]["Costo M.O. (Fijo)"]
        egreso_gastos_grales = df_egresos.iloc[i]["Gastos Grales (Fijos)"] 
        egreso_inversion = inversion_nov if i == 0 else 0
        otros_egresos = 0
        total_egreso = round(egreso_materiales + egreso_mano_obra + egreso_gastos_grales + egreso_inversion + otros_egresos, 2)
        
        saldo_fin_mes = round(saldo_mes_anterior + total_ingreso - total_egreso, 2)
        
        flujo_data.append({"DETALLES": "Efectivo al inicio del mes", month: saldo_mes_anterior})
        flujo_data.append({"DETALLES": "INGRESO DE EFECTIVO", month: None})
        flujo_data.append({"DETALLES": "Ingreso de efectivo de las ventas", month: ingreso_ventas})
        flujo_data.append({"DETALLES": "Ingreso de efectivo de las ventas a crédito", month: ingreso_credito})
        flujo_data.append({"DETALLES": "Otros ingresos de efectivo", month: otros_ingresos})
        flujo_data.append({"DETALLES": "TOTAL DE INGRESO DE EFECTIVO", month: total_ingreso})
        flujo_data.append({"DETALLES": "EGRESO DE EFECTIVO", month: None})
        flujo_data.append({"DETALLES": "Egreso de efectivo por Costos de Materiales", month: egreso_materiales})
        flujo_data.append({"DETALLES": "Egreso de efectivo por Costos de Mano de Obra", month: egreso_mano_obra})
        flujo_data.append({"DETALLES": "Egreso de efectivo por Gastos Generales", month: egreso_gastos_grales})
        flujo_data.append({"DETALLES": "Egreso de efectivo por inversión prevista en", month: egreso_inversion})
        flujo_data.append({"DETALLES": "Otros egresos de efectivo", month: otros_egresos})
        flujo_data.append({"DETALLES": "TOTAL DE EGRESO DE EFECTIVO", month: total_egreso})
        flujo_data.append({"DETALLES": "EFECTIVO AL FINAL DEL MES", month: saldo_fin_mes})
        
        saldo_mes_anterior = saldo_fin_mes

    # Formatear el DataFrame final
    df_flujo = pd.DataFrame(flujo_data).groupby("DETALLES").first().reindex([
        "Efectivo al inicio del mes", "INGRESO DE EFECTIVO", "Ingreso de efectivo de las ventas",
        "Ingreso de efectivo de las ventas a crédito", "Otros ingresos de efectivo", "TOTAL DE INGRESO DE EFECTIVO",
        "EGRESO DE EFECTIVO", "Egreso de efectivo por Costos de Materiales", "Egreso de efectivo por Costos de Mano de Obra",
        "Egreso de efectivo por Gastos Generales", "Egreso de efectivo por inversión prevista en",
        "Otros egresos de efectivo", "TOTAL DE EGRESO DE EFECTIVO", "EFECTIVO AL FINAL DEL MES"
    ])
    
    df_flujo["TOTAL"] = df_flujo.sum(axis=1)
    df_flujo.loc["Efectivo al inicio del mes", "TOTAL"] = None
    df_flujo.loc["EFECTIVO AL FINAL DEL MES", "TOTAL"] = round(df_flujo.iloc[-1]["Abr-26"], 2)

    return df_flujo, df_ingresos, df_egresos, bep_unidades, bep_soles

# --- Función para ejecutar la simulación de escenarios (Se mantiene) ---
def run_scenario(volumen_base, precio_base, crecimiento_base, cash_base, dias_cobro_base, scenario_type):
    
    if scenario_type == "Optimista":
        volumen_factor = 1.15
        precio_factor = 1.05
        crecimiento_factor = 1.10
        dias_cobro_factor = 0.8
    elif scenario_type == "Pesimista":
        volumen_factor = 0.85
        precio_factor = 0.95
        crecimiento_factor = 0.90
        dias_cobro_factor = 1.3
    else: 
        volumen_factor = precio_factor = crecimiento_factor = dias_cobro_factor = 1.0
        
    volumen_adj = volumen_base * volumen_factor
    precio_adj = precio_base * precio_factor
    crecimiento_adj = crecimiento_base * crecimiento_factor
    dias_cobro_adj = int(dias_cobro_base * dias_cobro_factor)
    
    df_flujo, _, df_egresos, _, _ = calcular_flujo_de_caja(
        volumen_inicial=volumen_adj,
        crecimiento_mensual=crecimiento_adj,
        ingreso_ecommerce=6000, 
        inversion_nov=2000, 
        cash_inicial=cash_base,
        precio_promedio=precio_adj,
        dias_cobro=dias_cobro_adj
    )
    
    saldo_final = df_flujo.loc["EFECTIVO AL FINAL DEL MES", "Abr-26"]
    utilidad_total = df_egresos["UTILIDAD NETA OPERACIONAL"].sum()
    
    return saldo_final, utilidad_total, precio_adj, volumen_adj


# --- 2. FUNCIÓN PARA CREAR EXCEL (No cambia) ---
def crear_bytes_excel(df_flujo_calc, df_ingresos_calc, df_egresos_calc):
    # (Se mantiene sin cambios)
    wb = openpyxl.Workbook()
    ws_flujo_caja = wb.active
    ws_flujo_caja.title = "Flujo de Caja OIT"

    currency_style = NamedStyle(name='currency_soles', number_format='"S/" #,##0.00')
    wb.add_named_style(currency_style)
    header_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    months = list(df_flujo_calc.columns[:-1]) 

    ws_flujo_caja.append(["DETALLES"] + months + ["TOTAL"])
    for cell in ws_flujo_caja[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for index, row in df_flujo_calc.iterrows():
        ws_flujo_caja.append([index] + list(row))
        
    for row in ws_flujo_caja.iter_rows(min_row=2, min_col=2, max_row=ws_flujo_caja.max_row, max_col=ws_flujo_caja.max_column):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.style = currency_style
    ws_flujo_caja.column_dimensions['A'].width = 40
    for i in range(len(months) + 1):
        col = get_column_letter(i + 2)
        ws_flujo_caja.column_dimensions[col].width = 18
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# --- 3. INTERFAZ DEL DASHBOARD (Streamlit) ---
st.title("🚗 Panel de Control y Proyección: Carwash 'El Profe'")
st.markdown("Análisis financiero avanzado.")

# --- Barra Lateral (Configuración) ---
st.sidebar.title("Gestión")
st.sidebar.markdown("---")
st.sidebar.header("⚙️ Panel de Escenarios")

# Sección de Eficiencia de Capital
st.sidebar.subheader("Eficiencia y Riesgo de Cobro")
dias_cobro_input = st.sidebar.slider(
    "1. Días Promedio de Cobro (DPC)",
    min_value=30, max_value=90, value=30, step=15,
    help="Días que tarda el cliente en pagar el 20% de la venta a crédito. Mayor DPC = Mayor riesgo de liquidez."
)

st.sidebar.subheader("Control de Ingresos")
precio_promedio_input = st.sidebar.slider(
    "2. Precio Promedio Ponderado (S/)",
    min_value=30.0, max_value=50.0, value=35.75, step=0.01,
    format="S/ %.2f",
    help="Define el precio promedio de venta de todos los servicios. El valor inicial S/ 35.75 asegura la precisión del primer mes."
)

# Sliders (mantienen las mejoras anteriores)
volumen_input = st.sidebar.slider(
    "3. Volumen Inicial de Servicios (Nov-25)",
    min_value=30, max_value=100, value=50, step=5,
    help="Número de autos que se espera lavar en el primer mes."
)
ecommerce_input = st.sidebar.slider(
    "4. Ingreso Fijo E-commerce (Mensual)",
    min_value=3000, max_value=10000, value=6000, step=500,
    help="Ingreso de la app y 'Club del Profe'."
)
st.sidebar.subheader("Control de Capital y Crecimiento")
cash_input = st.sidebar.number_input(
    "5. Efectivo Inicial (Capital de Trabajo)",
    min_value=0, max_value=20000, value=5000, step=1000
)
crecimiento_input = st.sidebar.slider(
    "6. Crecimiento Mensual de Servicios (%)",
    min_value=0.00, max_value=0.50, value=0.20, step=0.05,
    format="%.2f", 
    help="Porcentaje de crecimiento esperado del volumen de servicios."
)

# --- Calcular datos basados en los sliders ---
df_flujo, df_ingresos, df_egresos, bep_unidades, bep_soles = calcular_flujo_de_caja(
    volumen_inicial=volumen_input,
    crecimiento_mensual=crecimiento_input, 
    ingreso_ecommerce=ecommerce_input,
    inversion_nov=2000, 
    cash_inicial=cash_input,
    precio_promedio=precio_promedio_input,
    dias_cobro=dias_cobro_input 
)

# ----------------------------------------------------------------------
# EJECUCIÓN DEL ANÁLISIS DE ESCENARIOS (Mantenido para el Gráfico)
# ----------------------------------------------------------------------

scenarios = {}
scenarios["Base"] = run_scenario(volumen_input, precio_promedio_input, crecimiento_input, cash_input, dias_cobro_input, "Base")
scenarios["Optimista"] = run_scenario(volumen_input, precio_promedio_input, crecimiento_input, cash_input, dias_cobro_input, "Optimista")
scenarios["Pesimista"] = run_scenario(volumen_input, precio_promedio_input, crecimiento_input, cash_input, dias_cobro_input, "Pesimista")

df_escenarios = pd.DataFrame(
    data={
        "Escenario": ["Optimista", "Base", "Pesimista"],
        "Saldo_Final_Caja": [scenarios["Optimista"][0], scenarios["Base"][0], scenarios["Pesimista"][0]],
        "Utilidad_Neta": [scenarios["Optimista"][1], scenarios["Base"][1], scenarios["Pesimista"][1]],
        "Precio_Ajus": [scenarios["Optimista"][2], scenarios["Base"][2], scenarios["Pesimista"][2]],
        "Volumen_Ajus": [scenarios["Optimista"][3], scenarios["Base"][3], scenarios["Pesimista"][3]],
        "Color": ["#28A745", "#4682B4", "#DC3545"] 
    }
).set_index("Escenario")


# --- Sección de Scenarios (AHORA SOLO GRÁFICO AVANZADO) ---
st.header("🎯 Análisis de Escenarios")
st.markdown("Evaluación del proyecto bajo escenarios **Optimista**, **Base** y **Pesimista**.")

with st.container(border=True):
    col_chart, col_sensibilidad = st.columns([1.5, 1])
    
    # Gráfico de Riesgo (Visualización estilo Tornado Simplificado)
    chart_data = df_escenarios.reset_index()
    
    chart_riesgo = alt.Chart(chart_data).mark_bar().encode(
        x=alt.X('Saldo_Final_Caja', title='Saldo de Caja a 6 Meses (S/)'),
        y=alt.Y('Escenario', sort=['Optimista', 'Base', 'Pesimista']),
        color=alt.Color('Color', scale=alt.Scale(domain=chart_data['Color'].tolist(), range=chart_data['Color'].tolist())),
        tooltip=['Escenario', alt.Tooltip('Saldo_Final_Caja', format='$,.2f'), alt.Tooltip('Utilidad_Neta', format='$,.2f')]
    ).properties(
        title='Impacto en el Saldo de Caja Final'
    )
    
    col_chart.subheader("Gráfico de Impacto Financiero")
    col_chart.altair_chart(chart_riesgo, use_container_width=True)

    # Bloque de Resumen de Sensibilidad
    col_sensibilidad.subheader("Resumen de Sensibilidad")
    
    saldo_base = df_escenarios.loc['Base', 'Saldo_Final_Caja']
    saldo_pesimista = df_escenarios.loc['Pesimista', 'Saldo_Final_Caja']
    
    col_sensibilidad.metric("Saldo Final Base", f"S/ {saldo_base:,.2f}")
    col_sensibilidad.metric("Saldo Final Pesimista", f"S/ {saldo_pesimista:,.2f}",
                            delta=f"Pérdida de S/ {saldo_base - saldo_pesimista:,.2f} vs Base", delta_color="inverse")
                            
    col_sensibilidad.markdown(f"""
    <br>
    **Peor Caso (Pesimista):** <br>
    * Precio: S/ {df_escenarios.loc['Pesimista', 'Precio_Ajus']:.2f} 
    * Vol. Inicial: {df_escenarios.loc['Pesimista', 'Volumen_Ajus']:.0f} autos
    """, unsafe_allow_html=True)


# --- Análisis de la Opción Base (Continúa aquí) ---
st.header("📊 Análisis de la Opción Base")

with st.container(border=True):
    saldo_final = df_flujo.loc["EFECTIVO AL FINAL DEL MES", "Abr-26"]
    ingresos_totales = df_flujo.loc["TOTAL DE INGRESO DE EFECTIVO", "TOTAL"]
    total_egresos = df_flujo.loc["TOTAL DE EGRESO DE EFECTIVO", "TOTAL"]
    utilidad_total = df_egresos["UTILIDAD NETA OPERACIONAL"].sum()
    retorno_capital = (saldo_final - cash_input) / cash_input if cash_input > 0 else 0
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Ingresos Totales (6 Meses)", f"S/ {ingresos_totales:,.2f}")
    col2.metric("Egresos Totales (6 Meses)", f"S/ {total_egresos:,.2f}")
    col3.metric("Utilidad Neta Operacional", f"S/ {utilidad_total:,.2f}")
    col4.metric("Retorno de Capital (6M)", f"{retorno_capital*100:,.1f} %", delta=f"Inversión Inicial: S/ {cash_input:,.2f}")
    
    st.markdown("---")
    
    # NUEVA FILA DE MÉTRICAS AVANZADAS
    col_eficiencia, col_bep1, col_bep2, col_bep3 = st.columns([1, 1, 1, 1])
    col_eficiencia.metric("Eficiencia de Cobro (DPC)", f"{dias_cobro_input} días", 
                   help="Impacto de los días de cobro en la liquidez del Flujo de Caja.")
    
    # Análisis del Punto de Equilibrio (BEP)
    col_bep1.metric("Unidades BEP Mensual", f"{bep_unidades:,.0f} servicios", 
                   help="Cantidad mínima de servicios para cubrir costos fijos y variables mensuales.")
    col_bep2.metric("BEP Mensual (Soles)", f"S/ {bep_soles:,.2f}",
                   help="Monto mínimo de ingresos de servicios para cubrir costos fijos y variables mensuales.")

    vol_vs_bep_nov = df_ingresos.loc["Nov-25", "Volumen Servicios"] - bep_unidades
    col_bep3.metric("Margen de Seguridad (Nov-25)", f"{vol_vs_bep_nov:,.0f} unidades",
                   delta=f"Vol. Proy.: {df_ingresos.loc['Nov-25', 'Volumen Servicios']:,.0f} uds.")
                   
# -----------------------------------------------------------
# PASO 2: SEMÁFORO DE RIESGO DE LIQUIDEZ (Se mantiene)
# -----------------------------------------------------------
st.subheader("🚥 Semáforo de Riesgo de Liquidez")

saldo_minimo = df_flujo.loc["EFECTIVO AL FINAL DEL MES", :"Abr-26"].min()
cash_buffer = 1000 

if saldo_minimo <= 0:
    risk_level = "RIESGO CRÍTICO (Fondo negativo)"
    risk_class = "risk-high"
    risk_advice = f"🚨 El saldo de caja **cayó a S/ {saldo_minimo:,.2f}** o menos. ¡Cuidado! Un DPC alto puede ser la causa."
elif saldo_minimo < cash_buffer:
    risk_level = "RIESGO MEDIO (Buffer bajo)"
    risk_class = "risk-medium"
    risk_advice = f"⚠️ El saldo de caja mínimo es S/ {saldo_minimo:,.2f}. La tardanza en cobrar (DPC = {dias_cobro_input} días) está afectando la liquidez."
else:
    risk_level = "RIESGO BAJO (Posición Sólida)"
    risk_class = "risk-low"
    risk_advice = f"✅ El saldo de caja mínimo se mantiene en S/ {saldo_minimo:,.2f}. La posición es segura con DPC de {dias_cobro_input} días."

st.markdown(f'<div class="{risk_class}">{risk_advice}</div>', unsafe_allow_html=True)
st.markdown("---") 


# --- Gráficos Dinámicos (Se mantienen) ---
st.header("📉 Gráficos Dinámicos de Proyección")

with st.container(border=True):
    # SINCRONIZACIÓN: Filtro de meses 
    available_months = list(df_flujo.columns[:-1]) 
    selected_month_index = available_months.index("Abr-26") 

    filter_col, _ = st.columns([1, 3])
    end_month = filter_col.selectbox(
        "Selecciona el mes de corte para el análisis",
        options=available_months,
        index=selected_month_index,
        help="Elige hasta qué mes quieres visualizar los gráficos."
    )

    end_index = available_months.index(end_month) + 1
    months_to_show = available_months[:end_index]
    
    col_bar, col_line = st.columns(2)

    # 1. Gráfico de Barras: Ingresos vs Egresos
    chart_data_bar = pd.DataFrame({
        "Ingresos": df_flujo.loc["TOTAL DE INGRESO DE EFECTIVO", months_to_show],
        "Egresos": df_flujo.loc["TOTAL DE EGRESO DE EFECTIVO", months_to_show]
    })
    col_bar.subheader("Ingresos vs. Egresos Mensuales")
    col_bar.bar_chart(chart_data_bar, color=["#4682B4", "#A9A9A9"])

    # 2. Gráfico de Línea: Saldo de Caja
    chart_data_line = df_flujo.loc["EFECTIVO AL FINAL DEL MES", months_to_show]
    col_line.subheader("Evolución del Saldo de Caja")
    col_line.line_chart(chart_data_line, color="#4682B4")

    st.markdown("---")
    
    # NUEVA FILA DE GRÁFICOS: BEP y ANÁLISIS DE COSTOS
    col_bep_chart, col_costos_estructura, col_costos_detalles = st.columns([1.5, 1, 1])
    
    # 3. Gráfico de Rentabilidad vs. Punto de Equilibrio
    df_rentabilidad = pd.DataFrame({
        'Mes': months_to_show,
        'Volumen Proyectado': df_ingresos.loc[months_to_show, "Volumen Servicios"].values,
        'Punto de Equilibrio': [bep_unidades] * len(months_to_show)
    })
    
    chart_bep = alt.Chart(df_rentabilidad).encode(
        x=alt.X('Mes:O', sort=months_to_show),
        y=alt.Y('Volumen Proyectado', title='Volumen (Unidades)'),
        tooltip=['Mes', 'Volumen Proyectado', 'Punto de Equilibrio']
    ).properties(
        title='Rentabilidad de Volumen vs. Punto de Equilibrio (BEP)'
    )
    
    line_volumen = chart_bep.mark_line(point=True, color='#4682B4').encode(y='Volumen Proyectado')
    line_bep = chart_bep.mark_rule(color='red', strokeDash=[5, 5]).encode(y='Punto de Equilibrio', size=alt.value(2))
    final_chart = (line_volumen + line_bep).interactive()
    
    col_bep_chart.subheader("Análisis Gráfico de Cobertura de BEP")
    col_bep_chart.altair_chart(final_chart, use_container_width=True)
    
    # 4. Estructura de Costos Fijos vs. Variables
    costo_fijo_mes = df_egresos.loc[end_month, "Costo M.O. (Fijo)"] + df_egresos.loc[end_month, "Gastos Grales (Fijos)"]
    costo_variable_mes = df_egresos.loc[end_month, "Costo Mat. (Var)"]
    
    df_costos = pd.DataFrame({
        'Tipo': ['Costos Fijos', 'Costos Variables'],
        'Monto': [costo_fijo_mes, costo_variable_mes]
    })
    
    col_costos_estructura.subheader(f"Estructura de Costos ({end_month})")
    base_costo = alt.Chart(df_costos).encode(theta=alt.Theta("Monto", stack=True))
    pie_costo = base_costo.mark_arc(outerRadius=60).encode(
        color=alt.Color("Tipo", legend=None),
        order=alt.Order("Monto", sort="descending"),
        tooltip=["Tipo", alt.Tooltip("Monto", format="$,.0f")]
    ).properties(height=200)

    text_costo = base_costo.mark_text(radius=80).encode(
        text=alt.Text("Monto", format="$,.0f"),
        order=alt.Order("Monto", sort="descending"),
        color=alt.value("black")
    )
    col_costos_estructura.altair_chart(pie_costo + text_costo)

    # 5. Desglose de Costos Fijos
    df_fijos_detalles = pd.DataFrame({
        'Detalle': ['Mano de Obra', 'Alquiler', 'Servicios', 'Marketing', 'Capacitación/Ecológica'],
        'Monto': [
            df_egresos.loc[end_month, "Costo M.O. (Fijo)"],
            df_egresos.loc[end_month, "Alquiler"],
            df_egresos.loc[end_month, "Servicios (Luz/Agua)"],
            df_egresos.loc[end_month, "Marketing"],
            df_egresos.loc[end_month, "Capacitación"] + df_egresos.loc[end_month, "Cert. Ecológica"]
        ]
    })
    
    col_costos_detalles.subheader(f"Desglose de Costos Fijos")
    base_fijo = alt.Chart(df_fijos_detalles).encode(theta=alt.Theta("Monto", stack=True))
    pie_fijo = base_fijo.mark_arc(outerRadius=60).encode(
        color=alt.Color("Detalle", legend=alt.Legend(title="Gastos Fijos")),
        order=alt.Order("Monto", sort="descending"),
        tooltip=["Detalle", alt.Tooltip("Monto", format="$,.0f")]
    ).properties(height=200)

    col_costos_detalles.altair_chart(pie_fijo)


# --- Mostrar el Plan de Flujo de Caja OIT ---
with st.container(border=True):
    st.header("✅ Plan de Flujo de Caja (Formato OIT)")
    st.dataframe(df_flujo.style.format("S/ {:,.2f}", na_rep="---"))
    
    st.subheader("📥 Descargar Proyección")
    excel_bytes = crear_bytes_excel(df_flujo, df_ingresos, df_egresos)
    st.download_button(
        label="Descargar Archivo Excel (.xlsx) de este Escenario",
        data=excel_bytes,
        file_name=f"Proyeccion_ElProfe_V{volumen_input}_E{ecommerce_input}_DPC{dias_cobro_input}.xlsx",
        mime="application/vnd.ms-excel"
    )

# --- INCLUIDO DE NUEVO Y SIEMPRE VISIBLE: Detalle de Ingresos y Egresos ---
st.header("📚 Detalle de Ingresos y Egresos (Fuente de Cálculos)")
with st.container(border=True):
    st.subheader("Detalle de Ingresos Proyectados")
    st.dataframe(df_ingresos.style.format("S/ {:,.2f}", subset=pd.IndexSlice[:, "Ingr. Básico (Aprox)":]))

    st.subheader("Detalle de Egresos Proyectados")
    st.dataframe(df_egresos.style.format("S/ {:,.2f}"))